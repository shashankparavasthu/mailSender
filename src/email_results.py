from openpyxl import load_workbook
import config

def create_results_sheet():
    #create a new sheet to store the email results
    config.workbook.create_sheet(index = 1 , title = config.results_sheet_name) 
    # config.workbook.save(config.excel_file_path)

def add_cell_value(row_number, column_number, cell_value):

    #create cell object 
    cell = config.workbook[config.results_sheet_name].cell(row = row_number, column = column_number)
    cell.value = cell_value

def save_results():
    #save the workbook to retain the results
    config.workbook.save(config.excel_file_path) 
