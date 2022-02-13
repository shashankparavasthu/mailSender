from tkinter import filedialog
import email, smtplib, ssl
from openpyxl import load_workbook
import json 
import update_display

port = 465
context = ssl.create_default_context()
dynamic_character = '$'
extensions = ['.pdf', '.jpeg', '.png', '.jpg', '.JPG', '.PNG', '.JPEG', '.PDF']
results_sheet_name = "Email Results"

def init_recipients(mail_list_file):
    global excel_file_path
    global recipients
    global workbook
 
    excel_file_path = filedialog.askopenfilename(title="select excel file:")
    update_display.update_label(mail_list_file, excel_file_path)

    workbook = load_workbook(filename = excel_file_path)
    sheet = workbook.active

    recipients = {}
    for value in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        col_names = value
    
    recipient_count = 0
    for value in sheet.iter_rows(min_row=2, values_only=True):
        new_mail = {}
        recipient_count = recipient_count + 1
        for i in range(0,len(value)):
            new_mail[col_names[i]] = value[i]
        # recipients[value[0]] = new_mail
        recipients[str(recipient_count)] = new_mail

    print(json.dumps(recipients))


def set_attachment_folder(attachment_folder_label):
    global attachment_folder
    attachment_folder = filedialog.askdirectory(title="select attachment folder:")
    update_display.update_label(attachment_folder_label, attachment_folder)

def init_widgets():
    global widgets 
    widgets = []

def init_display(display_variable):
    global display_text
    display_text = display_variable
    
def init_input_screen(input_screen):
    global input_window
    input_window = input_screen

def add_widget(widget):
    widgets.append(widget)
